import re
import sqlite3
import os
import time
import json
import smtplib
from email.mime.text import MIMEText
from datetime import datetime

# Carrega .env o mais cedo possivel, antes de qualquer leitura de env var.
# Cobre todos os pontos de entrada (waitress app:app, run_prod.py, sync_forms.py).
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))
except ImportError:
    pass

from flask import Flask, render_template, request, jsonify, g, Response, redirect, url_for, flash, session
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_weasyprint import HTML

import ai  # assistente IA (OpenRouter) — lê config de env em runtime

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'estagios.db')

app = Flask(__name__)
app.config['DATABASE'] = DB_PATH

# SECRET_KEY: obrigatória em producao. Sem fallback inseguro hardcoded.
# Em desenvolvimento (FLASK_DEBUG=1) usa uma chave efemera so para nao travar o dev.
_secret = os.environ.get('SECRET_KEY')
if not _secret:
    if os.environ.get('FLASK_DEBUG') == '1':
        _secret = os.urandom(32).hex()
        print('[AVISO] SECRET_KEY ausente - usando chave efemera de desenvolvimento.')
    else:
        raise RuntimeError(
            'SECRET_KEY nao definida. Defina a variavel de ambiente SECRET_KEY '
            '(ex: no arquivo .env) antes de iniciar em producao.'
        )
app.secret_key = _secret

# ── Logging estruturado com rotacao ──────────────────────────────
# Grava em erp.log (5 MB x 5 arquivos). Essencial para auditoria/debug
# em producao, ja que o stdout do Waitress pode se perder.
import logging
from logging.handlers import RotatingFileHandler

_log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'erp.log')
_handler = RotatingFileHandler(_log_path, maxBytes=5 * 1024 * 1024, backupCount=5, encoding='utf-8')
_handler.setFormatter(logging.Formatter(
    '%(asctime)s %(levelname)s [%(remote_addr)s %(user)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
))


class _RequestContextFilter(logging.Filter):
    """Anexa IP e usuario logado a cada registro, quando houver request."""
    def filter(self, record):
        try:
            record.remote_addr = request.remote_addr or '-'
        except Exception:
            record.remote_addr = '-'
        try:
            record.user = current_user.username if current_user.is_authenticated else '-'
        except Exception:
            record.user = '-'
        return True


_handler.addFilter(_RequestContextFilter())
_handler.setLevel(logging.INFO)
app.logger.addHandler(_handler)
app.logger.setLevel(logging.INFO)

login_manager = LoginManager(app)
login_manager.login_view = 'login_page'

TIPO_ESTAGIO = {1: 'Observership', 2: 'Obrigatorio', 3: 'Optativo'}

ETAPAS_OBS = {
    1: 'Venda realizada',
    2: 'Pagamento confirmado',
    3: 'Docs enviados',
    4: 'Docs validados',
    5: 'Vaga confirmada',
    6: 'Orientacoes enviadas',
    7: 'Comprovante recebido',
    8: 'Concluido',
}

ETAPAS_OBR_OPT = {
    0: 'Verificacao de vaga',
    1: 'Venda realizada',
    2: 'Pagamento confirmado',
    3: 'Docs enviados',
    4: 'Docs validados',
    5: 'Vaga confirmada',
    6: 'Orientacoes enviadas',
    7: 'Comprovante recebido',
    8: 'Concluido',
}

MESES_PT = ['janeiro','fevereiro','março','abril','maio','junho',
            'julho','agosto','setembro','outubro','novembro','dezembro']

ETAPA_COLORS = {
    0: '#6b7280',
    1: '#f59e0b',
    2: '#3b82f6',
    3: '#8b5cf6',
    4: '#06b6d4',
    5: '#10b981',
    6: '#6366f1',
    7: '#0ea5e9',
    8: '#22c55e',
}

FORMAS_PAGAMENTO = ['PIX', 'Boleto', 'Cartao', 'Dinheiro', 'Isento', 'Outro']

TIPOS_RESIDENTE = ['Residente', 'Doutorando']
MODALIDADES_RESIDENTE = ['Optativo', 'Convenio']
STATUS_RESIDENTE = [
    'Interessado', 'Em andamento', 'Deferido', 'Confirmado',
    'Trocado', 'Indeferido', 'Desistente', 'Cancelado', 'Nao veio'
]
STATUS_RESIDENTE_COLORS = {
    'Interessado':  '#6b7280', 'Em andamento': '#f59e0b',
    'Deferido':     '#3b82f6', 'Confirmado':   '#10b981',
    'Trocado':      '#06b6d4', 'Indeferido':   '#dc2626',
    'Desistente':   '#9ca3af', 'Cancelado':    '#ef4444',
    'Nao veio':     '#374151',
}

MENSAGENS_MODELO_SEED = [
    ('whatsapp_aluno', 'Mensagem ao Aluno (WhatsApp)',
     'Olá {{nome}}, tudo bem?\n{{usuario}} aqui do Ensino e Pesquisa.',
     'nome (nome do aluno), usuario (primeiro nome de quem esta enviando)'),
    ('whatsapp_area_medica', 'Mensagem a Area Medica (WhatsApp)',
     'Olá! Tudo bem?\nGostaria de verificar a disponibilidade de vaga para Estágio {{modalidade}} em '
     '{{especialidade}} para o {{tipo}} {{nome}}, {{periodo}}.\nFico no aguardo do retorno. Muito obrigado!',
     'nome (nome do aluno), tipo (residente/doutorando), modalidade (Optativo/Convenio), '
     'especialidade (nome oficial do contato), periodo (frase ja formatada), '
     'usuario (primeiro nome de quem esta enviando)'),
]

ITENS_POR_PAGINA = 15


# ── Email config ──────────────────────────────────────────────
EMAIL_CONFIG = {
    'enabled': os.environ.get('SMTP_ENABLED', 'false').lower() == 'true',
    'host': os.environ.get('SMTP_HOST', 'smtp.gmail.com'),
    'port': int(os.environ.get('SMTP_PORT', '587')),
    'user': os.environ.get('SMTP_USER', ''),
    'pass': os.environ.get('SMTP_PASS', ''),
    'from_addr': os.environ.get('SMTP_FROM', ''),
}


def enviar_email(notificacao):
    """Envia email se SMTP estiver configurado. Registra notificacao de qualquer forma."""
    db = get_db()
    db.execute('''INSERT INTO notificacoes (estagio_id, tipo, mensagem, email_destino, enviado)
                  VALUES (?, ?, ?, ?, ?)''',
               (notificacao['estagio_id'], notificacao['tipo'], notificacao['mensagem'],
                notificacao['email'], EMAIL_CONFIG['enabled']))
    db.commit()

    if not EMAIL_CONFIG['enabled'] or not notificacao['email']:
        return

    try:
        msg = MIMEText(notificacao['mensagem'], 'plain', 'utf-8')
        msg['Subject'] = notificacao['assunto']
        msg['From'] = EMAIL_CONFIG['from_addr'] or EMAIL_CONFIG['user']
        msg['To'] = notificacao['email']

        with smtplib.SMTP(EMAIL_CONFIG['host'], EMAIL_CONFIG['port']) as server:
            server.starttls()
            server.login(EMAIL_CONFIG['user'], EMAIL_CONFIG['pass'])
            server.send_message(msg)

        db.execute('UPDATE notificacoes SET enviado=1 WHERE estagio_id=? AND tipo=? ORDER BY id DESC LIMIT 1',
                   (notificacao['estagio_id'], notificacao['tipo']))
        db.commit()
    except Exception as e:
        app.logger.warning(f'Falha ao enviar email: {e}')


# ── Auth ──────────────────────────────────────────────────────
class User(UserMixin):
    def __init__(self, id_, username, nome, role):
        self.id = id_
        self.username = username
        self.nome = nome
        self.role = role


@login_manager.user_loader
def load_user(user_id):
    db = get_db()
    row = db.execute('SELECT * FROM usuarios WHERE id=?', (user_id,)).fetchone()
    if not row:
        return None
    return User(row['id'], row['username'], row['nome'], row['role'])


# ── Database ──────────────────────────────────────────────────
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(app.config['DATABASE'], timeout=10)
        g.db.row_factory = sqlite3.Row
        g.db.execute('PRAGMA foreign_keys = ON')
        # WAL: permite leituras concorrentes com escrita; essencial com
        # Waitress rodando multiplas threads sobre o mesmo arquivo SQLite.
        g.db.execute('PRAGMA journal_mode = WAL')
        g.db.execute('PRAGMA synchronous = NORMAL')
        g.db.execute('PRAGMA busy_timeout = 5000')
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def hash_password(password):
    """Simple password hashing using hashlib (no bcrypt dependency)."""
    import hashlib
    salt = os.urandom(16).hex()
    h = hashlib.sha256((salt + password).encode()).hexdigest()
    return f'{salt}:{h}'


def verify_password(password, stored):
    """Verify password against stored hash."""
    import hashlib
    if ':' not in stored:
        return False
    salt, h = stored.split(':', 1)
    return hashlib.sha256((salt + password).encode()).hexdigest() == h


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.execute('PRAGMA foreign_keys = OFF')
    db.executescript('''
        CREATE TABLE IF NOT EXISTS tipo_estagio (
            id INTEGER PRIMARY KEY,
            nome TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            nome TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user'
        );

        CREATE TABLE IF NOT EXISTS estagios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo_id INTEGER NOT NULL REFERENCES tipo_estagio(id),
            mes_ano TEXT NOT NULL,
            semana INTEGER NOT NULL,
            nome TEXT NOT NULL,
            cpf TEXT,
            especialidade TEXT NOT NULL,
            cracha TEXT,
            valor REAL,
            forma_pagamento TEXT,
            status_pagamento TEXT DEFAULT 'Pendente',
            comprovante_pagamento TEXT,
            inicio DATE,
            termino DATE,
            email TEXT,
            telefone TEXT,
            observacao TEXT,
            documentos TEXT,
            envio_certificado DATE,
            etapa INTEGER NOT NULL DEFAULT 0,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS historico_etapas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            estagio_id INTEGER NOT NULL REFERENCES estagios(id),
            etapa INTEGER NOT NULL,
            observacao TEXT,
            responsavel TEXT,
            ts DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS notificacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            estagio_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            mensagem TEXT NOT NULL,
            email_destino TEXT,
            enviado INTEGER DEFAULT 0,
            ts DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS limite_especialidade (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            especialidade TEXT UNIQUE NOT NULL COLLATE NOCASE,
            limite_semanal INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS residentes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            email TEXT,
            telefone TEXT,
            cpf TEXT,
            tipo TEXT NOT NULL DEFAULT 'Residente',
            modalidade TEXT NOT NULL DEFAULT 'Optativo',
            especialidade TEXT NOT NULL,
            subespecialidade TEXT,
            instituicao_origem TEXT,
            programa_ano TEXT,
            mes_ano TEXT NOT NULL,
            inicio DATE,
            termino DATE,
            status TEXT NOT NULL DEFAULT 'Interessado',
            valor REAL,
            forma_pagamento TEXT,
            status_pagamento TEXT DEFAULT 'Pendente',
            comprovante_pagamento TEXT,
            observacao TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS historico_residentes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            residente_id INTEGER NOT NULL REFERENCES residentes(id),
            status TEXT NOT NULL,
            observacao TEXT,
            responsavel TEXT,
            ts DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS area_medica (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            especialidade TEXT NOT NULL,
            nome TEXT,
            celular TEXT,
            email TEXT,
            obs_internato TEXT,
            obs_residencia TEXT
        );

        CREATE TABLE IF NOT EXISTS mensagens_modelo (
            chave TEXT PRIMARY KEY,
            titulo TEXT NOT NULL,
            texto TEXT NOT NULL,
            placeholders TEXT
        );

        DELETE FROM historico_etapas;
        DELETE FROM notificacoes;
        DELETE FROM estagios;
        DELETE FROM tipo_estagio;
        DELETE FROM usuarios;

        INSERT INTO tipo_estagio (id, nome) VALUES (1, 'Observership');
        INSERT INTO tipo_estagio (id, nome) VALUES (2, 'Obrigatorio');
        INSERT INTO tipo_estagio (id, nome) VALUES (3, 'Optativo');

    ''')
    db.commit()

    # Insert default users with proper password hashes
    admin_hash = hash_password('admin')
    user_hash = hash_password('user')
    db.execute("INSERT INTO usuarios (username, password_hash, nome, role) VALUES (?, ?, ?, ?)",
               ('admin', admin_hash, 'Administrador', 'admin'))
    db.execute("INSERT INTO usuarios (username, password_hash, nome, role) VALUES (?, ?, ?, ?)",
               ('user', user_hash, 'Usuario', 'user'))
    db.executemany(
        'INSERT OR IGNORE INTO mensagens_modelo (chave, titulo, texto, placeholders) VALUES (?,?,?,?)',
        MENSAGENS_MODELO_SEED
    )
    db.commit()

    count = db.execute('SELECT COUNT(*) FROM estagios').fetchone()[0]
    if count == 0:
        db.executescript('''
            INSERT INTO estagios (id, tipo_id, mes_ano, semana, nome, cpf, especialidade, cracha, valor, forma_pagamento, status_pagamento, inicio, termino, email, telefone, observacao, documentos, envio_certificado, etapa)
            VALUES
            (1, 1, '2025-06', 1, 'Ana Silva', '123.456.789-00', 'Cardiologia', 'OBS-001', 1500.00, 'PIX', 'Pago', '2025-06-01', '2025-07-15', 'ana@email.com', '(51) 99999-0001', 'Aluna do programa de observership', 'CRM;Termo', '2025-06-10', 2),
            (2, 2, '2025-06', 2, 'Bruno Costa', '987.654.321-00', 'Cirurgia Geral', 'OBR-002', 500, 'PIX', 'Pago', '2025-06-01', '2025-12-31', 'bruno@email.com', '(51) 99999-0002', 'Estagio obrigatorio 6o periodo', 'CRM;Vacina', NULL, 0),
            (3, 3, '2025-06', 2, 'Carla Souza', '456.789.123-00', 'Pediatria', 'OPT-003', 800.00, 'Boleto', 'Pendente', '2025-06-01', '2025-08-30', 'carla@email.com', '(51) 99999-0003', 'Estagio optativo de pediatria', 'CRM', NULL, 1);

            INSERT INTO historico_etapas (estagio_id, etapa, observacao, responsavel)
            VALUES
            (1, 1, 'Venda registrada', 'admin'),
            (1, 2, 'Pagamento via PIX confirmado', 'admin'),
            (2, 0, 'Solicitacao de vaga enviada a Cirurgia', 'admin'),
            (3, 0, 'Vaga confirmada pela Pediatria', 'admin'),
            (3, 1, 'Venda registrada', 'user');
        ''')
        db.commit()

    db.execute('PRAGMA foreign_keys = ON')
    db.close()


# ── Auth routes ───────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'POST':
        data = request.form if request.form else request.get_json()
        username = data.get('username', '')
        password = data.get('password', '')
        db = get_db()
        row = db.execute('SELECT * FROM usuarios WHERE username=?', (username,)).fetchone()
        if row and verify_password(password, row['password_hash']):
            user = User(row['id'], row['username'], row['nome'], row['role'])
            login_user(user)
            db.execute('UPDATE usuarios SET last_login=CURRENT_TIMESTAMP WHERE id=?', (user.id,))
            db.commit()
            app.logger.info(f'Login OK: {username}')
            next_page = request.args.get('next', '/residentes')
            if request.accept_mimetypes.accept_json and not request.accept_mimetypes.accept_html:
                return jsonify({'ok': True, 'nome': user.nome, 'role': user.role})
            return redirect(next_page)
        app.logger.warning(f'Login FALHOU: usuario="{username}"')
        if request.accept_mimetypes.accept_json and not request.accept_mimetypes.accept_html:
            return jsonify({'erro': 'Credenciais invalidas'}), 401
        return render_template('login.html', erro='Usuario ou senha invalidos'), 401
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login_page'))


# ── Pages ─────────────────────────────────────────────────────
@app.route('/health')
def health():
    """Health check sem autenticacao para monitoramento/auto-restart.
    Verifica que o processo responde e o banco esta acessivel."""
    try:
        get_db().execute('SELECT 1')
        return jsonify({'status': 'ok'}), 200
    except Exception as e:
        return jsonify({'status': 'error', 'detalhe': str(e)}), 503


@app.route('/')
@login_required
def index():
    # Modulo Observership desativado no ERP: fluxo agora e gerenciado pela e-commerce.
    # Dados historicos continuam no banco (usados por dashboard/relatorios/IA),
    # mas a tela de CRUD nao fica mais acessivel via menu/URL.
    return redirect('/residentes')


@app.route('/dashboard')
@login_required
def dashboard_page():
    return render_template('dashboard.html')


@app.route('/relatorios')
@login_required
def relatorios_page():
    return render_template('relatorios.html')


# ── API: Auth ─────────────────────────────────────────────────
@app.route('/api/me')
@login_required
def api_me():
    return jsonify({'id': current_user.id, 'username': current_user.username,
                    'nome': current_user.nome, 'role': current_user.role})


# ── API: Estagios ─────────────────────────────────────────────
@app.route('/api/estagios', methods=['GET'])
@login_required
def api_get_estagios():
    db = get_db()
    query = '''
        SELECT e.*, t.nome as tipo_nome,
            CAST(julianday('now') - julianday(
                COALESCE(
                    (SELECT MAX(h.ts) FROM historico_etapas h WHERE h.estagio_id = e.id),
                    e.updated_at, e.created_at
                )
            ) AS INTEGER) as dias_na_etapa
        FROM estagios e
        JOIN tipo_estagio t ON e.tipo_id = t.id
        WHERE 1=1
    '''
    params = []

    tipo_id = request.args.get('tipo_id')
    if tipo_id:
        query += ' AND e.tipo_id = ?'
        params.append(tipo_id)

    especialidade = request.args.get('especialidade')
    if especialidade:
        query += ' AND e.especialidade = ?'
        params.append(especialidade)

    etapa = request.args.get('etapa')
    if etapa is not None and etapa != '':
        query += ' AND e.etapa = ?'
        params.append(etapa)

    mes_ano = request.args.get('mes_ano')
    if mes_ano:
        query += ' AND e.mes_ano = ?'
        params.append(mes_ano)

    busca = request.args.get('busca')
    if busca:
        query += ' AND (e.nome LIKE ? OR e.email LIKE ? OR e.cracha LIKE ? OR e.cpf LIKE ? OR e.especialidade LIKE ? OR e.observacao LIKE ?)'
        params.extend([f'%{busca}%'] * 6)

    status_pag = request.args.get('status_pagamento')
    if status_pag:
        query += ' AND e.status_pagamento = ?'
        params.append(status_pag)

    semana = request.args.get('semana')
    if semana:
        query += ' AND e.semana = ?'
        params.append(int(semana))

    # Count total for pagination
    count_query = query.replace('SELECT e.*, t.nome as tipo_nome', 'SELECT COUNT(*)')
    total = db.execute(count_query, params).fetchone()[0]

    # Pagination
    page = max(1, int(request.args.get('page', 1)))
    per_page = min(100, max(1, int(request.args.get('per_page', ITENS_POR_PAGINA))))
    offset = (page - 1) * per_page
    query += ' ORDER BY e.mes_ano DESC, e.semana, e.nome LIMIT ? OFFSET ?'
    params.extend([per_page, offset])

    rows = db.execute(query, params).fetchall()
    result = []
    for r in rows:
        result.append({
            'id': r['id'],
            'tipo_id': r['tipo_id'],
            'tipo_nome': r['tipo_nome'],
            'mes_ano': r['mes_ano'],
            'semana': r['semana'],
            'nome': r['nome'],
            'cpf': r['cpf'],
            'especialidade': r['especialidade'],
            'cracha': r['cracha'],
            'valor': r['valor'],
            'forma_pagamento': r['forma_pagamento'],
            'status_pagamento': r['status_pagamento'],
            'comprovante_pagamento': r['comprovante_pagamento'],
            'inicio': r['inicio'],
            'termino': r['termino'],
            'email': r['email'],
            'telefone': r['telefone'],
            'observacao': r['observacao'],
            'documentos': r['documentos'],
            'envio_certificado': r['envio_certificado'],
            'etapa': r['etapa'],
            'dias_na_etapa': r['dias_na_etapa'] or 0,
            'created_at': r['created_at'],
            'updated_at': r['updated_at'],
        })
    return jsonify({
        'data': result,
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': max(1, (total + per_page - 1) // per_page)
    })


@app.route('/api/estagios', methods=['POST'])
@login_required
def api_create_estagio():
    db = get_db()
    data = request.get_json()
    tipo_id = data.get('tipo_id')
    if tipo_id in (2, 3):
        etapa_inicial = 0
    else:
        etapa_inicial = 1

    try:
        cursor = db.execute('''
            INSERT INTO estagios (tipo_id, mes_ano, semana, nome, cpf, especialidade, cracha,
                valor, forma_pagamento, status_pagamento, comprovante_pagamento,
                inicio, termino, email, telefone, observacao, documentos,
                envio_certificado, etapa)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            tipo_id, data.get('mes_ano'), data.get('semana'), data.get('nome'),
            data.get('cpf'), data.get('especialidade'), data.get('cracha'),
            data.get('valor'), data.get('forma_pagamento'),
            data.get('status_pagamento', 'Interessado'), data.get('comprovante_pagamento'),
            data.get('inicio'), data.get('termino'), data.get('email'), data.get('telefone'),
            data.get('observacao'), data.get('documentos'), data.get('envio_certificado'),
            etapa_inicial,
        ))
        estagio_id = cursor.lastrowid
    except sqlite3.IntegrityError as e:
        if 'cracha' in str(e).lower() or 'UNIQUE' in str(e):
            return jsonify({'erro': 'Cracha ja cadastrado para outro estagio'}), 400
        raise

    responsavel = current_user.nome if current_user.is_authenticated else 'Sistema'
    db.execute('''
        INSERT INTO historico_etapas (estagio_id, etapa, observacao, responsavel)
        VALUES (?, ?, ?, ?)
    ''', (estagio_id, etapa_inicial, 'Registro criado', responsavel))
    db.commit()
    return jsonify({'id': estagio_id, 'etapa': etapa_inicial}), 201


@app.route('/api/estagios/<int:estagio_id>', methods=['PUT'])
@login_required
def api_update_estagio(estagio_id):
    db = get_db()
    existing = db.execute('SELECT * FROM estagios WHERE id=?', (estagio_id,)).fetchone()
    if not existing:
        return jsonify({'erro': 'Estagio nao encontrado'}), 404
    data = request.get_json()
    # Merge: use sent values, fall back to existing
    fields = ['tipo_id','mes_ano','semana','nome','cpf','especialidade','cracha',
              'valor','forma_pagamento','status_pagamento','comprovante_pagamento',
              'inicio','termino','email','telefone','observacao','documentos',
              'envio_certificado','carga_horaria','comprovante_estagio']
    vals = {f: data.get(f, existing[f]) for f in fields}
    try:
        db.execute('''
            UPDATE estagios SET
                tipo_id=?, mes_ano=?, semana=?, nome=?, cpf=?, especialidade=?,
                cracha=?, valor=?, forma_pagamento=?, status_pagamento=?,
                comprovante_pagamento=?, inicio=?, termino=?, email=?, telefone=?,
                observacao=?, documentos=?, envio_certificado=?,
                carga_horaria=?, comprovante_estagio=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        ''', (
            vals['tipo_id'], vals['mes_ano'], vals['semana'],
            vals['nome'], vals['cpf'], vals['especialidade'],
            vals['cracha'], vals['valor'], vals['forma_pagamento'],
            vals['status_pagamento'], vals['comprovante_pagamento'],
            vals['inicio'], vals['termino'], vals['email'],
            vals['telefone'], vals['observacao'], vals['documentos'],
            vals['envio_certificado'], vals['carga_horaria'],
            vals['comprovante_estagio'], estagio_id,
        ))
        db.commit()
    except sqlite3.IntegrityError as e:
        if 'UNIQUE' in str(e):
            return jsonify({'erro': 'Cracha ja cadastrado para outro estagio'}), 400
        raise
    return jsonify({'ok': True})


@app.route('/api/estagios/<int:estagio_id>', methods=['DELETE'])
@login_required
def api_delete_estagio(estagio_id):
    db = get_db()
    db.execute('DELETE FROM notificacoes WHERE estagio_id=?', (estagio_id,))
    db.execute('DELETE FROM historico_etapas WHERE estagio_id=?', (estagio_id,))
    db.execute('DELETE FROM estagios WHERE id=?', (estagio_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/estagios/<int:estagio_id>/avancar', methods=['POST'])
@login_required
def api_avancar_etapa(estagio_id):
    db = get_db()
    row = db.execute('SELECT e.*, t.nome as tipo_nome FROM estagios e JOIN tipo_estagio t ON e.tipo_id = t.id WHERE e.id=?', (estagio_id,)).fetchone()
    if not row:
        return jsonify({'erro': 'Estagio nao encontrado'}), 404

    etapa_atual = row['etapa']
    tipo_id = row['tipo_id']
    max_etapa = 8
    if tipo_id == 1:
        min_etapa = 1
    else:
        min_etapa = 0

    if etapa_atual >= max_etapa:
        return jsonify({'erro': 'Estagio ja concluido'}), 400

    nova_etapa = etapa_atual + 1
    data = request.get_json() or {}
    responsavel = data.get('responsavel') or (current_user.nome if current_user.is_authenticated else 'Sistema')

    extra = ", status_pagamento='Pago'" if nova_etapa == 2 else ""
    db.execute(f'UPDATE estagios SET etapa=?{extra}, updated_at=CURRENT_TIMESTAMP WHERE id=?', (nova_etapa, estagio_id))
    db.execute('''
        INSERT INTO historico_etapas (estagio_id, etapa, observacao, responsavel)
        VALUES (?, ?, ?, ?)
    ''', (estagio_id, nova_etapa, data.get('observacao', ''), responsavel))
    db.commit()

    # Send email notification
    if row['email']:
        etapa_nome = (ETAPAS_OBS if tipo_id == 1 else ETAPAS_OBR_OPT).get(nova_etapa, '')
        enviar_email({
            'estagio_id': estagio_id,
            'tipo': 'avanco_etapa',
            'assunto': f'Estagio atualizado - {etapa_nome}',
            'mensagem': f'Ola {row["nome"]},\n\nSeu estagio em {row["especialidade"]} avancou para a etapa: {nova_etapa} - {etapa_nome}.\nResponsavel: {responsavel}\n\nSanta Casa / UFCSPA',
            'email': row['email'],
        })

    return jsonify({'etapa': nova_etapa})


@app.route('/api/estagios/<int:estagio_id>/historico', methods=['GET'])
@login_required
def api_historico(estagio_id):
    db = get_db()
    rows = db.execute('SELECT * FROM historico_etapas WHERE estagio_id=? ORDER BY ts', (estagio_id,)).fetchall()
    result = []
    for r in rows:
        result.append({
            'id': r['id'],
            'estagio_id': r['estagio_id'],
            'etapa': r['etapa'],
            'observacao': r['observacao'],
            'responsavel': r['responsavel'],
            'ts': r['ts'],
        })
    return jsonify(result)


@app.route('/api/estagios/<int:estagio_id>/pdf')
@login_required
def api_estagio_pdf(estagio_id):
    db = get_db()
    row = db.execute('SELECT e.*, t.nome as tipo_nome FROM estagios e JOIN tipo_estagio t ON e.tipo_id = t.id WHERE e.id=?', (estagio_id,)).fetchone()
    if not row:
        return jsonify({'erro': 'Nao encontrado'}), 404

    historico = db.execute('SELECT * FROM historico_etapas WHERE estagio_id=? ORDER BY ts', (estagio_id,)).fetchall()

    etapa_nome = (ETAPAS_OBS if row['tipo_id'] == 1 else ETAPAS_OBR_OPT).get(row['etapa'], '')

    html = render_template('pdf_ficha.html', estagio=row, historico=historico, etapa_nome=etapa_nome, data_geracao=datetime.now().strftime('%d/%m/%Y %H:%M'))
    try:
        pdf_bytes = HTML(string=html).write_pdf()
        return Response(pdf_bytes, mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename=estagio_{estagio_id}.pdf'})
    except Exception:
        # Fallback: render as printable HTML
        return html


# ── API: Certificado PDF ──────────────────────────────────────
@app.route('/api/estagios/<int:estagio_id>/certificado')
@login_required
def api_estagio_certificado(estagio_id):
    db = get_db()
    row = db.execute('SELECT e.*, t.nome as tipo_nome FROM estagios e JOIN tipo_estagio t ON e.tipo_id = t.id WHERE e.id=?', (estagio_id,)).fetchone()
    if not row:
        return jsonify({'erro': 'Nao encontrado'}), 404
    if row['etapa'] < 7:
        return jsonify({'erro': 'Comprovante de estagio nao recebido (etapa minima: 7)'}), 400
    now = datetime.now()
    data_ext = f"Porto Alegre, {now.day:02d} de {MESES_PT[now.month - 1]} de {now.year}"

    def fmt_data(s):
        if not s:
            return '—'
        parts = str(s).split('T')[0].split('-')
        return f"{parts[2]}/{parts[1]}/{parts[0]}" if len(parts) == 3 else s

    html = render_template('certificado.html', estagio=row, data_geracao=data_ext,
                           inicio_fmt=fmt_data(row['inicio']), termino_fmt=fmt_data(row['termino']))
    try:
        pdf_bytes = HTML(string=html).write_pdf()
        nome_arquivo = row['nome'].replace(' ', '_')[:40]
        return Response(pdf_bytes, mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename=certificado_{nome_arquivo}.pdf'})
    except Exception:
        return html


# ── API: Marcar Pago ──────────────────────────────────────────
@app.route('/api/estagios/<int:estagio_id>/pago', methods=['POST'])
@login_required
def api_marcar_pago(estagio_id):
    db = get_db()
    if not db.execute('SELECT id FROM estagios WHERE id=?', (estagio_id,)).fetchone():
        return jsonify({'erro': 'Nao encontrado'}), 404
    db.execute("UPDATE estagios SET status_pagamento='Pago', updated_at=CURRENT_TIMESTAMP WHERE id=?", (estagio_id,))
    db.commit()
    return jsonify({'ok': True})


# ── API: Dropdowns ────────────────────────────────────────────
@app.route('/api/tipos', methods=['GET'])
@login_required
def api_tipos():
    db = get_db()
    # ?em_uso=1 retorna apenas tipos com pelo menos 1 estagio cadastrado —
    # usado para simplificar o filtro de listagem (evita mostrar tipos
    # sem nenhum registro, ex: Obrigatorio/Optativo hoje nao utilizados
    # no modulo Estagios; esse fluxo passou a acontecer via modulo Residentes).
    if request.args.get('em_uso') == '1':
        rows = db.execute('''
            SELECT t.* FROM tipo_estagio t
            WHERE EXISTS (SELECT 1 FROM estagios e WHERE e.tipo_id = t.id)
            ORDER BY t.id
        ''').fetchall()
    else:
        rows = db.execute('SELECT * FROM tipo_estagio ORDER BY id').fetchall()
    return jsonify([{'id': r['id'], 'nome': r['nome']} for r in rows])


@app.route('/api/especialidades', methods=['GET'])
@login_required
def api_especialidades():
    db = get_db()
    rows = db.execute('SELECT DISTINCT especialidade FROM estagios ORDER BY especialidade').fetchall()
    return jsonify([r['especialidade'] for r in rows])


@app.route('/api/meses', methods=['GET'])
@login_required
def api_meses():
    db = get_db()
    rows = db.execute('SELECT DISTINCT mes_ano FROM estagios ORDER BY mes_ano DESC').fetchall()
    return jsonify([r['mes_ano'] for r in rows])


@app.route('/api/formas-pagamento', methods=['GET'])
@login_required
def api_formas_pagamento():
    return jsonify(FORMAS_PAGAMENTO)


# ── API: Dashboard ────────────────────────────────────────────
@app.route('/api/dashboard')
@login_required
def api_dashboard():
    db = get_db()
    mes_filtro = request.args.get('mes_ano', '').strip()  # filtro opcional
    where_mes = " WHERE mes_ano=?" if mes_filtro else ""
    params_mes = [mes_filtro] if mes_filtro else []

    # ── Residentes & Doutorandos (unico foco do dashboard) ──
    res_total = db.execute(
        f"SELECT COUNT(*) FROM residentes{where_mes}", params_mes
    ).fetchone()[0]

    res_por_status = db.execute(
        f"SELECT status, COUNT(*) as cnt FROM residentes{where_mes} "
        f"GROUP BY status ORDER BY cnt DESC", params_mes
    ).fetchall()

    res_por_tipo = db.execute(
        f"SELECT COALESCE(tipo,'N/A') as tipo, COUNT(*) as cnt FROM residentes"
        f"{where_mes} GROUP BY tipo ORDER BY cnt DESC", params_mes
    ).fetchall()

    res_por_especialidade = db.execute(
        f"SELECT COALESCE(especialidade,'N/A') as esp, COUNT(*) as cnt FROM residentes"
        f"{where_mes} GROUP BY especialidade ORDER BY cnt DESC LIMIT 15", params_mes
    ).fetchall()

    res_financeiro = db.execute(
        f"SELECT COALESCE(SUM(valor),0) as total, "
        f"COALESCE(SUM(CASE WHEN status_pagamento='Pago' THEN valor ELSE 0 END),0) as pago "
        f"FROM residentes{where_mes}", params_mes
    ).fetchone()

    # Tendência mensal — últimos 18 meses (sem filtro de mês, sempre geral)
    res_por_mes = db.execute('''
        SELECT mes_ano, COUNT(*) as cnt FROM residentes
        WHERE mes_ano IS NOT NULL AND mes_ano != ''
        GROUP BY mes_ano ORDER BY mes_ano DESC LIMIT 18
    ''').fetchall()

    # KPIs de status + alertas de "dias parado" (mesma logica da badge da lista)
    res_kpis = db.execute(f'''
        SELECT
            COUNT(*) FILTER (WHERE status='Interessado') as novos,
            COUNT(*) FILTER (WHERE status='Em andamento') as em_andamento,
            COUNT(*) FILTER (WHERE status='Deferido') as deferidos,
            COUNT(*) FILTER (WHERE status='Confirmado') as confirmados,
            COUNT(*) FILTER (WHERE status_pagamento='Pendente'
                AND status NOT IN ('Cancelado','Indeferido','Desistente','Nao veio')) as pag_pendente,
            COUNT(*) FILTER (WHERE status IN ('Interessado','Em andamento','Deferido') AND dias > 14) as criticos,
            COUNT(*) FILTER (WHERE status IN ('Interessado','Em andamento','Deferido')
                AND dias > 7 AND dias <= 14) as alertas
        FROM (
            SELECT r.status, r.status_pagamento,
                CAST(julianday('now') - julianday(COALESCE(
                    (SELECT MAX(h.ts) FROM historico_residentes h WHERE h.residente_id = r.id),
                    r.updated_at, r.created_at)) AS INTEGER) AS dias
            FROM residentes r{where_mes}
        )
    ''', params_mes).fetchone()

    recent = db.execute(f'''
        SELECT id, nome, status, tipo, especialidade, status_pagamento, updated_at
        FROM residentes{where_mes}
        ORDER BY updated_at DESC LIMIT 8
    ''', params_mes).fetchall()

    # Lista de meses disponíveis para o filtro
    meses_disp = [r[0] for r in db.execute('''
        SELECT DISTINCT mes_ano FROM residentes
        WHERE mes_ano IS NOT NULL AND mes_ano != ''
        ORDER BY mes_ano DESC
    ''').fetchall()]

    return jsonify({
        'total': res_total,
        'por_status': [{'status': r['status'], 'count': r['cnt']} for r in res_por_status],
        'por_tipo': [{'tipo': r['tipo'], 'count': r['cnt']} for r in res_por_tipo],
        'por_especialidade': [{'nome': r['esp'], 'count': r['cnt']} for r in res_por_especialidade],
        'por_mes': [{'mes_ano': r['mes_ano'], 'count': r['cnt']} for r in res_por_mes],
        'financeiro': {
            'total': res_financeiro['total'] or 0,
            'pago': res_financeiro['pago'] or 0,
            'pendente': (res_financeiro['total'] or 0) - (res_financeiro['pago'] or 0),
        },
        'kpis': {
            'novos': res_kpis['novos'] or 0,
            'em_andamento': res_kpis['em_andamento'] or 0,
            'deferidos': res_kpis['deferidos'] or 0,
            'confirmados': res_kpis['confirmados'] or 0,
            'pag_pendente': res_kpis['pag_pendente'] or 0,
            'criticos': res_kpis['criticos'] or 0,
            'alertas': res_kpis['alertas'] or 0,
        },
        'recentes': [{
            'id': r['id'], 'nome': r['nome'], 'status': r['status'], 'tipo': r['tipo'],
            'especialidade': r['especialidade'], 'status_pagamento': r['status_pagamento'],
            'updated_at': r['updated_at']
        } for r in recent],
        'meses_disponiveis': meses_disp,
        'mes_filtro': mes_filtro or None,
    })


# ── API: Pendencias ──────────────────────────────────────────
@app.route('/api/pendencias', methods=['GET'])
@login_required
def api_pendencias():
    db = get_db()
    rows = db.execute('''
        SELECT
            COUNT(*) FILTER (WHERE etapa < 7)                                  AS em_andamento,
            COUNT(*) FILTER (WHERE etapa < 7 AND dias > 14)                    AS criticos,
            COUNT(*) FILTER (WHERE etapa < 7 AND dias > 7 AND dias <= 14)      AS alertas,
            COUNT(*) FILTER (WHERE status_pagamento = "Pendente" AND etapa < 7) AS pag_pendente
        FROM (
            SELECT e.etapa, e.status_pagamento,
                CAST(julianday("now") - julianday(
                    COALESCE(
                        (SELECT MAX(h.ts) FROM historico_etapas h WHERE h.estagio_id = e.id),
                        e.updated_at, e.created_at
                    )
                ) AS INTEGER) AS dias
            FROM estagios e
            WHERE e.tipo_id = 1
        )
    ''').fetchone()

    # Pendências de Residentes & Doutorandos
    res_rows = db.execute('''
        SELECT
            COUNT(*) FILTER (WHERE status = "Interessado")                       AS novos,
            COUNT(*) FILTER (WHERE status = "Em andamento")                      AS em_andamento,
            COUNT(*) FILTER (WHERE status = "Deferido")                          AS deferidos,
            COUNT(*) FILTER (WHERE status = "Confirmado")                        AS confirmados,
            COUNT(*) FILTER (WHERE status_pagamento = "Pendente"
                             AND status NOT IN ("Cancelado","Indeferido","Desistente","Nao veio")) AS pag_pendente
        FROM residentes
    ''').fetchone()

    return jsonify({
        'em_andamento':       rows['em_andamento']          or 0,
        'criticos':           rows['criticos']              or 0,
        'alertas':            rows['alertas']               or 0,
        'pag_pendente':       rows['pag_pendente']          or 0,
        'res_novos':          res_rows['novos']             or 0,
        'res_em_andamento':   res_rows['em_andamento']      or 0,
        'res_deferidos':      res_rows['deferidos']         or 0,
        'res_confirmados':    res_rows['confirmados']       or 0,
        'res_pag_pendente':   res_rows['pag_pendente']      or 0,
    })


# ── API: Assistente IA (OpenRouter) ────────────────────────────
@app.route('/api/ai/status', methods=['GET'])
@login_required
def api_ai_status():
    return jsonify({'enabled': ai.is_enabled()})


@app.route('/api/ai/chat', methods=['POST'])
@login_required
def api_ai_chat():
    if not ai.is_enabled():
        return jsonify({'erro': 'Assistente de IA não está habilitado.'}), 503
    data = request.get_json(silent=True) or {}
    historico = data.get('messages') or []
    if not isinstance(historico, list) or not historico:
        return jsonify({'erro': 'Nenhuma mensagem enviada.'}), 400
    # Limita o histórico para conter custo/contexto.
    historico = historico[-12:]
    try:
        snapshot = ai.montar_snapshot(get_db())
        mensagens = ai.montar_mensagens(snapshot, historico)
        inicio = time.time()
        resposta = ai.chamar_openrouter(mensagens)
        ms = int((time.time() - inicio) * 1000)
        ultima = (historico[-1].get('content') or '')[:120]
        app.logger.info(f'IA chat OK ({ms}ms): "{ultima}"')
        return jsonify({'resposta': resposta})
    except ai.AIError as e:
        app.logger.warning(f'IA chat FALHOU: {e}')
        return jsonify({'erro': str(e)}), 502


@app.route('/api/ai/insights', methods=['GET'])
@login_required
def api_ai_insights():
    if not ai.is_enabled():
        return jsonify({'erro': 'Assistente de IA não está habilitado.'}), 503
    try:
        snapshot = ai.montar_snapshot(get_db())
        pedido = {
            'role': 'user',
            'content': (
                'Faça um resumo executivo curto da situação atual de Residentes & '
                'Doutorandos. Destaque o que precisa de atenção: novas inscrições '
                '(Interessado), pagamentos pendentes e gargalos por status. '
                'Use no máximo 8 linhas, em tópicos.'
            ),
        }
        mensagens = ai.montar_mensagens(snapshot, [pedido])
        inicio = time.time()
        resumo = ai.chamar_openrouter(mensagens, temperature=0.2, max_tokens=2000)
        ms = int((time.time() - inicio) * 1000)
        app.logger.info(f'IA insights OK ({ms}ms)')
        return jsonify({'resumo': resumo})
    except ai.AIError as e:
        app.logger.warning(f'IA insights FALHOU: {e}')
        return jsonify({'erro': str(e)}), 502


# ── API: CSV Export ───────────────────────────────────────────
@app.route('/api/exportar-csv', methods=['GET'])
@login_required
def api_exportar_csv():
    db = get_db()
    query = '''
        SELECT e.*, t.nome as tipo_nome
        FROM estagios e
        JOIN tipo_estagio t ON e.tipo_id = t.id
        WHERE 1=1
    '''
    params = []

    tipo_id = request.args.get('tipo_id')
    if tipo_id:
        query += ' AND e.tipo_id = ?'
        params.append(tipo_id)

    especialidade = request.args.get('especialidade')
    if especialidade:
        query += ' AND e.especialidade = ?'
        params.append(especialidade)

    etapa = request.args.get('etapa')
    if etapa is not None and etapa != '':
        query += ' AND e.etapa = ?'
        params.append(etapa)

    mes_ano = request.args.get('mes_ano')
    if mes_ano:
        query += ' AND e.mes_ano = ?'
        params.append(mes_ano)

    busca = request.args.get('busca')
    if busca:
        query += ' AND (e.nome LIKE ? OR e.email LIKE ? OR e.cracha LIKE ? OR e.cpf LIKE ? OR e.especialidade LIKE ? OR e.observacao LIKE ?)'
        params.extend([f'%{busca}%'] * 6)

    query += ' ORDER BY e.mes_ano DESC, e.semana, e.nome'

    rows = db.execute(query, params).fetchall()

    sep = request.args.get('separador', ';')
    if sep not in (';', ',', '\t'):
        sep = ';'

    def get_etapa_nome(tipo_id, etapa):
        if tipo_id == 1:
            return ETAPAS_OBS.get(etapa, '')
        return ETAPAS_OBR_OPT.get(etapa, '')

    lines = [
        sep.join(['ID', 'Tipo', 'Mes/Ano', 'Semana', 'Nome', 'CPF', 'Especialidade',
                   'Cracha', 'Valor', 'Forma Pagto', 'Status Pagto', 'Inicio',
                   'Termino', 'Email', 'Telefone', 'Documentos', 'Observacao',
                   'Certificado', 'Etapa'])
    ]
    for r in rows:
        etapa_nome = get_etapa_nome(r['tipo_id'], r['etapa'])
        lines.append(sep.join(str(v) if v is not None else '' for v in [
            r['id'], r['tipo_nome'], r['mes_ano'], r['semana'],
            r['nome'], r['cpf'], r['especialidade'], r['cracha'], r['valor'],
            r['forma_pagamento'], r['status_pagamento'], r['inicio'],
            r['termino'], r['email'], r['telefone'], r['documentos'],
            r['observacao'], r['envio_certificado'], f"{r['etapa']} - {etapa_nome}"
        ]))

    csv_content = '\r\n'.join(lines)
    ext = 'tsv' if sep == '\t' else 'csv'
    mime = 'text/tab-separated-values' if sep == '\t' else 'text/csv'
    resp = Response(
        '﻿' + csv_content,
        mimetype=mime,
        headers={'Content-Disposition': f'attachment; filename=estagios.{ext}'}
    )
    resp.headers['Content-Type'] = f'{mime}; charset=utf-8'
    return resp


# ── API: Relatórios ──────────────────────────────────────────
@app.route('/api/relatorios/exportar', methods=['GET'])
@login_required
def api_relatorios_exportar():
    db = get_db()
    relatorio = request.args.get('relatorio', 'egressos')
    especialidade = request.args.get('especialidade', '')
    tipo_id = request.args.get('tipo_id', '')
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')
    mes_ano = request.args.get('mes_ano', '')
    etapa_filtro = request.args.get('etapa', '')
    dias_min = int(request.args.get('dias_min', 7) or 7)
    preview = request.args.get('preview', '')

    sql = '''SELECT e.*, t.nome as tipo_nome,
             CAST(julianday('now') - julianday(
                 COALESCE((SELECT MAX(h.ts) FROM historico_etapas h WHERE h.estagio_id=e.id), e.created_at)
             ) AS INTEGER) as dias_na_etapa
             FROM estagios e JOIN tipo_estagio t ON e.tipo_id=t.id WHERE 1=1'''
    params = []

    if especialidade:
        sql += ' AND e.especialidade = ?'
        params.append(especialidade)
    if tipo_id:
        sql += ' AND e.tipo_id = ?'
        params.append(tipo_id)

    if relatorio == 'egressos':
        sql += ' AND e.etapa = 8'
        if data_inicio:
            sql += ' AND e.termino >= ?'
            params.append(data_inicio)
        if data_fim:
            sql += ' AND e.termino <= ?'
            params.append(data_fim)
        sql += ' ORDER BY e.termino DESC, e.nome'
        headers = ['Nome', 'Email', 'Telefone', 'Especialidade', 'Tipo', 'Inicio', 'Termino', 'Carga Horaria (h)']
        def row_cols(r):
            return [r['nome'], r['email'], r['telefone'], r['especialidade'],
                    r['tipo_nome'], r['inicio'], r['termino'], r['carga_horaria']]

    elif relatorio == 'pendentes':
        sql += " AND e.status_pagamento IN ('Pendente','Interessado') AND e.etapa < 8"
        if etapa_filtro != '':
            sql += ' AND e.etapa = ?'
            params.append(etapa_filtro)
        sql += ' ORDER BY e.mes_ano DESC, e.nome'
        headers = ['Nome', 'Email', 'Telefone', 'Especialidade', 'Tipo', 'Valor', 'Status Pagto', 'Etapa']
        def row_cols(r):
            etapa_nome = (ETAPAS_OBS if r['tipo_id'] == 1 else ETAPAS_OBR_OPT).get(r['etapa'], '')
            return [r['nome'], r['email'], r['telefone'], r['especialidade'],
                    r['tipo_nome'], r['valor'], r['status_pagamento'],
                    f"{r['etapa']} - {etapa_nome}"]

    elif relatorio == 'parados':
        sql += ' AND e.etapa < 8 AND e.etapa > 0'
        sql += ' ORDER BY e.especialidade, e.nome'
        headers = ['Nome', 'Email', 'Telefone', 'Especialidade', 'Tipo', 'Etapa', 'Dias na Etapa']
        def row_cols(r):
            etapa_nome = (ETAPAS_OBS if r['tipo_id'] == 1 else ETAPAS_OBR_OPT).get(r['etapa'], '')
            return [r['nome'], r['email'], r['telefone'], r['especialidade'],
                    r['tipo_nome'], f"{r['etapa']} - {etapa_nome}", r['dias_na_etapa']]

    else:  # especialidade
        if mes_ano:
            sql += ' AND e.mes_ano = ?'
            params.append(mes_ano)
        sql += ' ORDER BY e.especialidade, e.nome'
        headers = ['Especialidade', 'Nome', 'Email', 'Telefone', 'Tipo', 'Mes/Ano', 'Inicio', 'Termino']
        def row_cols(r):
            return [r['especialidade'], r['nome'], r['email'], r['telefone'],
                    r['tipo_nome'], r['mes_ano'], r['inicio'], r['termino']]

    rows = db.execute(sql, params).fetchall()

    if relatorio == 'parados':
        rows = [r for r in rows if (r['dias_na_etapa'] or 0) >= dias_min]

    if preview:
        data = []
        for r in rows[:50]:
            cols = row_cols(r)
            data.append(dict(zip(headers, [str(v) if v is not None else '' for v in cols])))
        return jsonify({'data': data, 'total': len(rows), 'headers': headers})

    sep = ';'
    lines = [sep.join(headers)]
    for r in rows:
        lines.append(sep.join(str(v) if v is not None else '' for v in row_cols(r)))
    csv_content = '\r\n'.join(lines)
    nome_arquivo = f'relatorio_{relatorio}.csv'
    resp = Response(
        '﻿' + csv_content,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={nome_arquivo}'}
    )
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    return resp


# ── API: Notificacoes ─────────────────────────────────────────
@app.route('/api/notificacoes', methods=['GET'])
@login_required
def api_notificacoes():
    db = get_db()
    estagio_id = request.args.get('estagio_id')
    if estagio_id:
        rows = db.execute('SELECT * FROM notificacoes WHERE estagio_id=? ORDER BY ts DESC', (estagio_id,)).fetchall()
    else:
        rows = db.execute('SELECT * FROM notificacoes ORDER BY ts DESC LIMIT 50').fetchall()
    return jsonify([{
        'id': r['id'], 'estagio_id': r['estagio_id'], 'tipo': r['tipo'],
        'mensagem': r['mensagem'], 'email_destino': r['email_destino'],
        'enviado': r['enviado'], 'ts': r['ts']
    } for r in rows])


# ── API: Importar Excel ───────────────────────────────────────
_MESES_PT = {
    'janeiro': 1, 'fevereiro': 2, 'março': 3, 'marco': 3,
    'abril': 4, 'maio': 5, 'junho': 6, 'julho': 7,
    'agosto': 8, 'setembro': 9, 'outubro': 10,
    'novembro': 11, 'dezembro': 12,
}
_ABAS_IGNORAR = {
    'especialidade x vagas', 'lista de presença', 'lista de presenca',
    'certificados estágios', 'certificados estagios',
    'semana feriado', 'crachás', 'crachas', 'lista de interessados',
    'lista de presença ',
}
_COL_IMPORT = {
    'nome': 'nome', 'nome ': 'nome',
    'especialidade': 'especialidade', 'especialidade ': 'especialidade',
    'crachá': 'cracha', 'cracha': 'cracha', 'crachá ': 'cracha',
    'valor': 'valor', 'valor ': 'valor',
    'término': 'termino', 'termino': 'termino', 'término ': 'termino',
    'e-mail': 'email', 'email': 'email', 'e-mail ': 'email',
    'telefone': 'telefone', 'telefone ': 'telefone',
    'observação': 'observacao', 'observacao': 'observacao', 'observação ': 'observacao',
    'envio de certificado': 'envio_certificado', 'envio de certificado ': 'envio_certificado',
    'etapa certificado': 'envio_certificado', 'etapa certificado ': 'envio_certificado',
    'documentação': 'documentos', 'documentos': 'documentos',
    'documentação ': 'documentos', 'documentos ': 'documentos',
    'documentacao': 'documentos', 'documentação': 'documentos',
    'modalidade': '_modalidade', 'modalidade ': '_modalidade',
}


def _xclean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return None if v in ('\xa0', '', '-', 'N/A') else v
    return v


def _xvalor(v):
    if v is None or v == '\xa0':
        return None
    if isinstance(v, (int, float)):
        return float(v) if v else None
    if isinstance(v, str):
        v = re.sub(r'[R$\xa0\s]', '', v).replace('.', '').replace(',', '.')
        try:
            return float(v) or None
        except ValueError:
            return None
    return None


def _xdate(v):
    if v is None:
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    return None


def _parse_aba_xlsx(ws, mes_ano_fixo, ano):
    records = []
    col_map = {}
    semana = 1
    semana_from_date = False
    current_mes_ano = mes_ano_fixo

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if all(c is None or (isinstance(c, str) and c.strip() in ('', '\xa0')) for c in row):
            continue

        col0 = row[0]

        if isinstance(col0, str):
            col0s = col0.strip()
            col0l = col0s.lower()

            # Cabeçalho de mês (ex: "Janeiro") — usado na aba 2026
            if col0l.rstrip() in _MESES_PT:
                mes_num = _MESES_PT[col0l.rstrip()]
                current_mes_ano = f'{ano}-{mes_num:02d}'
                col_map = {}
                semana = 1
                semana_from_date = False
                continue

            # Linha de cabeçalho de semana com número: "Semana 1 | Nome | ..."
            m = re.match(r'semana\s*(\d+)', col0l)
            if m:
                semana = int(m.group(1))
                semana_from_date = False
                new_map = {}
                for j, cell in enumerate(row):
                    if j == 0 or cell is None:
                        continue
                    key = str(cell).strip().lower()
                    if key in _COL_IMPORT:
                        new_map[j] = _COL_IMPORT[key]
                if new_map:
                    col_map = new_map
                continue

            # "Semana " sem número (aba 2026, meses a partir de Março)
            if re.match(r'^semana\s*$', col0l):
                semana_from_date = True
                new_map = {}
                for j, cell in enumerate(row):
                    if j == 0 or cell is None:
                        continue
                    key = str(cell).strip().lower()
                    if key in _COL_IMPORT:
                        new_map[j] = _COL_IMPORT[key]
                if new_map:
                    col_map = new_map
                continue
            continue

        # Linha de dados: col0 deve ser uma data
        if not hasattr(col0, 'strftime') or not col_map:
            continue

        # Inferir semana a partir do dia quando header não traz número
        if semana_from_date:
            semana = (col0.day - 1) // 7 + 1

        # Extrair nome
        nome = None
        for j, field in col_map.items():
            if field == 'nome' and j < len(row):
                nome = _xclean(row[j])
                break
        if not nome:
            continue

        rec = {
            'mes_ano': current_mes_ano,
            'semana': semana,
            'nome': nome,
            'tipo_id': 1,
        }

        for j, field in col_map.items():
            if j >= len(row):
                continue
            v = row[j]
            if field == 'nome':
                pass
            elif field == 'valor':
                rec['valor'] = _xvalor(v)
            elif field == 'termino':
                rec['termino'] = _xdate(v)
            elif field == 'envio_certificado':
                cv = _xclean(v)
                rec['envio_certificado'] = _xdate(v) if hasattr(v, 'strftime') else None
            elif field in ('especialidade', 'email', 'telefone', 'observacao', 'documentos'):
                rec[field] = _xclean(v)
            elif field == 'cracha':
                cv = _xclean(v)
                if not cv or str(cv) in ('0', '0.0') or (isinstance(cv, str) and cv.lower() == 'devolvido'):
                    rec['cracha'] = None
                elif str(cv).replace('.', '').isdigit():
                    rec['cracha'] = str(int(float(str(cv))))
                else:
                    rec['cracha'] = str(cv)
            elif field == '_modalidade':
                cv = (_xclean(v) or '').lower()
                if 'obrig' in cv:
                    rec['tipo_id'] = 2
                elif 'opta' in cv:
                    rec['tipo_id'] = 3

        if not rec.get('especialidade'):
            continue

        rec['etapa'] = 1 if rec['tipo_id'] == 1 else 0
        records.append(rec)

    return records


@app.route('/api/importar-excel', methods=['POST'])
@login_required
def api_importar_excel():
    if 'arquivo' not in request.files:
        return jsonify({'erro': 'Nenhum arquivo enviado'}), 400
    arquivo = request.files['arquivo']
    if not arquivo.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'erro': 'Arquivo deve ser .xlsx ou .xls'}), 400

    confirmar = request.args.get('confirmar', '0') == '1'
    # Filtro de abas: lista de nomes selecionados pelo usuário (case-insensitive)
    abas_selecionadas = request.form.getlist('abas') or []
    abas_selecionadas_lower = {a.strip().lower() for a in abas_selecionadas}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    except Exception as e:
        return jsonify({'erro': f'Erro ao ler arquivo: {e}'}), 400

    # Listar abas válidas (não ignoradas) para retornar no preview
    abas_validas = []
    for sn in wb.sheetnames:
        nl = sn.strip().lower()
        if nl in _ABAS_IGNORAR:
            continue
        if nl == '2026':
            abas_validas.append(sn)
        else:
            parts = re.split(r'[-/]', nl)
            if len(parts) >= 2 and parts[0].strip() in _MESES_PT:
                abas_validas.append(sn)

    all_records = []
    for sheet_name in abas_validas:
        name_lower = sheet_name.strip().lower()
        # Aplicar filtro de seleção (se houver)
        if abas_selecionadas_lower and name_lower not in abas_selecionadas_lower:
            continue

        if name_lower == '2026':
            ano, mes_ano_fixo = 2026, None
        else:
            parts = re.split(r'[-/]', name_lower)
            mes_nome = parts[0].strip()
            ano_str = parts[-1].strip()
            try:
                ano = int(ano_str) if len(ano_str) == 4 else 2000 + int(ano_str)
            except ValueError:
                continue
            mes = _MESES_PT[mes_nome]
            mes_ano_fixo = f'{ano}-{mes:02d}'

        ws = wb[sheet_name]
        all_records.extend(_parse_aba_xlsx(ws, mes_ano_fixo, ano))

    # Dedup contra registros existentes (lower() em Python pois SQLite não trata acentos)
    db = get_db()
    existing = set()
    for row in db.execute('SELECT nome, coalesce(especialidade,""), mes_ano FROM estagios').fetchall():
        existing.add((row[0].lower() if row[0] else '', row[1].lower(), row[2]))

    novos, duplicados = [], []
    for rec in all_records:
        key = (rec['nome'].lower(), (rec.get('especialidade') or '').lower(), rec.get('mes_ano', ''))
        if key in existing:
            duplicados.append(rec)
        else:
            novos.append(rec)
            existing.add(key)

    if not confirmar:
        return jsonify({
            'total_planilha': len(all_records),
            'novos': len(novos),
            'duplicados': len(duplicados),
            'preview': novos[:30],
            'abas_disponiveis': abas_validas,
        })

    responsavel = current_user.nome if current_user.is_authenticated else 'Importacao'
    importados = 0
    for rec in novos:
        try:
            cursor = db.execute('''
                INSERT INTO estagios (tipo_id, mes_ano, semana, nome, especialidade, cracha,
                    valor, termino, email, telefone, observacao, documentos,
                    envio_certificado, etapa)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                rec.get('tipo_id', 1), rec.get('mes_ano'), rec.get('semana', 1),
                rec['nome'], rec.get('especialidade'), rec.get('cracha'),
                rec.get('valor'), rec.get('termino'), rec.get('email'),
                rec.get('telefone'), rec.get('observacao'), rec.get('documentos'),
                rec.get('envio_certificado'), rec.get('etapa', 1),
            ))
            db.execute('''
                INSERT INTO historico_etapas (estagio_id, etapa, observacao, responsavel)
                VALUES (?, ?, ?, ?)
            ''', (cursor.lastrowid, rec.get('etapa', 1), 'Importado via planilha Excel', responsavel))
            importados += 1
        except Exception:
            pass
    db.commit()

    return jsonify({
        'total_planilha': len(all_records),
        'importados': importados,
        'duplicados': len(duplicados),
    })


# ── Admin: Usuarios ──────────────────────────────────────────
@app.route('/usuarios')
@login_required
def pagina_usuarios():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    return render_template('usuarios.html')


@app.route('/api/usuarios', methods=['GET'])
@login_required
def api_listar_usuarios():
    if current_user.role != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403
    db = get_db()
    rows = db.execute('SELECT id, username, nome, role, last_login FROM usuarios ORDER BY id').fetchall()
    return jsonify([{
        'id': r['id'], 'username': r['username'], 'nome': r['nome'],
        'role': r['role'], 'last_login': r['last_login']
    } for r in rows])


@app.route('/api/usuarios', methods=['POST'])
@login_required
def api_criar_usuario():
    if current_user.role != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403
    db = get_db()
    data = request.get_json()
    username = data.get('username', '').strip()
    nome = data.get('nome', '').strip()
    senha = data.get('senha', '')
    role = data.get('role', 'user')
    if not username or not nome or not senha:
        return jsonify({'erro': 'Preencha todos os campos obrigatorios'}), 400
    existing = db.execute('SELECT id FROM usuarios WHERE username=?', (username,)).fetchone()
    if existing:
        return jsonify({'erro': 'Username ja existe'}), 409
    pw_hash = hash_password(senha)
    db.execute('INSERT INTO usuarios (username, password_hash, nome, role) VALUES (?,?,?,?)',
               (username, pw_hash, nome, role))
    db.commit()
    uid = db.execute('SELECT last_insert_rowid()').fetchone()[0]
    return jsonify({'id': uid, 'username': username, 'nome': nome, 'role': role}), 201


@app.route('/api/usuarios/<int:user_id>', methods=['PUT'])
@login_required
def api_editar_usuario(user_id):
    if current_user.role != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403
    db = get_db()
    data = request.get_json()
    existing = db.execute('SELECT * FROM usuarios WHERE id=?', (user_id,)).fetchone()
    if not existing:
        return jsonify({'erro': 'Usuario nao encontrado'}), 404
    nome = data.get('nome', existing['nome'])
    role = data.get('role', existing['role'])
    username = data.get('username', existing['username'])
    # Check username conflict
    dup = db.execute('SELECT id FROM usuarios WHERE username=? AND id!=?', (username, user_id)).fetchone()
    if dup:
        return jsonify({'erro': 'Username ja existe'}), 409
    senha = data.get('senha', '')
    if senha:
        pw_hash = hash_password(senha)
        db.execute('UPDATE usuarios SET username=?, nome=?, role=?, password_hash=? WHERE id=?',
                   (username, nome, role, pw_hash, user_id))
    else:
        db.execute('UPDATE usuarios SET username=?, nome=?, role=? WHERE id=?',
                   (username, nome, role, user_id))
    db.commit()
    return jsonify({'id': user_id, 'username': username, 'nome': nome, 'role': role})


@app.route('/api/usuarios/<int:user_id>', methods=['DELETE'])
@login_required
def api_excluir_usuario(user_id):
    if current_user.role != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403
    db = get_db()
    if user_id == current_user.id:
        return jsonify({'erro': 'Nao pode excluir o proprio usuario'}), 400
    existing = db.execute('SELECT * FROM usuarios WHERE id=?', (user_id,)).fetchone()
    if not existing:
        return jsonify({'erro': 'Usuario nao encontrado'}), 404
    db.execute('DELETE FROM usuarios WHERE id=?', (user_id,))
    db.commit()
    return jsonify({'ok': True})


# ── Admin: Vagas (limites por especialidade) ─────────────────
@app.route('/vagas')
@login_required
def pagina_vagas():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    return render_template('vagas.html')


@app.route('/api/limites', methods=['GET'])
@login_required
def api_listar_limites():
    if current_user.role != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403
    db = get_db()
    rows = db.execute('SELECT id, especialidade, limite_semanal FROM limite_especialidade ORDER BY especialidade COLLATE NOCASE').fetchall()
    return jsonify([{'id': r['id'], 'especialidade': r['especialidade'], 'limite_semanal': r['limite_semanal']} for r in rows])


@app.route('/api/limites', methods=['POST'])
@login_required
def api_criar_limite():
    if current_user.role != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403
    db = get_db()
    data = request.get_json()
    especialidade = (data.get('especialidade') or '').strip()
    limite = data.get('limite_semanal')
    if not especialidade or limite is None:
        return jsonify({'erro': 'Preencha especialidade e limite'}), 400
    try:
        db.execute('INSERT INTO limite_especialidade (especialidade, limite_semanal) VALUES (?,?)', (especialidade, int(limite)))
        db.commit()
        row_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]
        return jsonify({'id': row_id, 'especialidade': especialidade, 'limite_semanal': int(limite)}), 201
    except sqlite3.IntegrityError:
        return jsonify({'erro': 'Especialidade ja cadastrada'}), 409


@app.route('/api/limites/<int:limite_id>', methods=['PUT'])
@login_required
def api_editar_limite(limite_id):
    if current_user.role != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403
    db = get_db()
    data = request.get_json()
    row = db.execute('SELECT * FROM limite_especialidade WHERE id=?', (limite_id,)).fetchone()
    if not row:
        return jsonify({'erro': 'Nao encontrado'}), 404
    especialidade = (data.get('especialidade') or row['especialidade']).strip()
    limite = int(data.get('limite_semanal', row['limite_semanal']))
    try:
        db.execute('UPDATE limite_especialidade SET especialidade=?, limite_semanal=? WHERE id=?', (especialidade, limite, limite_id))
        db.commit()
        return jsonify({'id': limite_id, 'especialidade': especialidade, 'limite_semanal': limite})
    except sqlite3.IntegrityError:
        return jsonify({'erro': 'Especialidade ja cadastrada'}), 409


@app.route('/api/limites/<int:limite_id>', methods=['DELETE'])
@login_required
def api_excluir_limite(limite_id):
    if current_user.role != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403
    db = get_db()
    if not db.execute('SELECT id FROM limite_especialidade WHERE id=?', (limite_id,)).fetchone():
        return jsonify({'erro': 'Nao encontrado'}), 404
    db.execute('DELETE FROM limite_especialidade WHERE id=?', (limite_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/vagas-semana', methods=['GET'])
@login_required
def api_vagas_semana():
    mes_ano = request.args.get('mes_ano', '')
    semana = request.args.get('semana', '')
    if not mes_ano or not semana:
        return jsonify({'erro': 'mes_ano e semana obrigatorios'}), 400
    db = get_db()
    rows = db.execute('''
        SELECT l.id, l.especialidade, l.limite_semanal,
               COALESCE(c.usadas, 0) as usadas
        FROM limite_especialidade l
        LEFT JOIN (
            SELECT especialidade, COUNT(*) as usadas
            FROM estagios
            WHERE mes_ano=? AND semana=?
            GROUP BY especialidade
        ) c ON lower(l.especialidade) = lower(c.especialidade)
        ORDER BY l.especialidade COLLATE NOCASE
    ''', (mes_ano, int(semana))).fetchall()
    return jsonify([{
        'especialidade': r['especialidade'],
        'limite': r['limite_semanal'],
        'usadas': r['usadas'],
        'livres': max(0, r['limite_semanal'] - r['usadas']),
    } for r in rows])


@app.route('/configuracoes')
@login_required
def pagina_configuracoes():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    return render_template('configuracoes.html')


# ── Contatos da Area Medica (chefes de servico por especialidade) ──
@app.route('/api/area-medica', methods=['GET'])
@login_required
def api_area_medica():
    db = get_db()
    rows = db.execute('SELECT * FROM area_medica ORDER BY especialidade COLLATE NOCASE').fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/area-medica', methods=['POST'])
@login_required
def api_area_medica_criar():
    if current_user.role != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403
    d = request.get_json() or {}
    especialidade = (d.get('especialidade') or '').strip()
    if not especialidade:
        return jsonify({'erro': 'Especialidade e obrigatoria'}), 400
    db = get_db()
    cur = db.execute(
        '''INSERT INTO area_medica (especialidade, nome, celular, email, obs_internato, obs_residencia)
           VALUES (?,?,?,?,?,?)''',
        (especialidade, d.get('nome', ''), d.get('celular', ''), d.get('email', ''),
         d.get('obs_internato', ''), d.get('obs_residencia', ''))
    )
    db.commit()
    return jsonify({'id': cur.lastrowid}), 201


@app.route('/api/area-medica/<int:cid>', methods=['PUT'])
@login_required
def api_area_medica_editar(cid):
    if current_user.role != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403
    if not get_db().execute('SELECT id FROM area_medica WHERE id=?', (cid,)).fetchone():
        return jsonify({'erro': 'Nao encontrado'}), 404
    d = request.get_json() or {}
    especialidade = (d.get('especialidade') or '').strip()
    if not especialidade:
        return jsonify({'erro': 'Especialidade e obrigatoria'}), 400
    db = get_db()
    db.execute(
        '''UPDATE area_medica SET especialidade=?, nome=?, celular=?, email=?, obs_internato=?, obs_residencia=?
           WHERE id=?''',
        (especialidade, d.get('nome', ''), d.get('celular', ''), d.get('email', ''),
         d.get('obs_internato', ''), d.get('obs_residencia', ''), cid)
    )
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/area-medica/<int:cid>', methods=['DELETE'])
@login_required
def api_area_medica_excluir(cid):
    if current_user.role != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403
    db = get_db()
    if not db.execute('SELECT id FROM area_medica WHERE id=?', (cid,)).fetchone():
        return jsonify({'erro': 'Nao encontrado'}), 404
    db.execute('DELETE FROM area_medica WHERE id=?', (cid,))
    db.commit()
    return jsonify({'ok': True})


# ── Modelos de mensagem (WhatsApp aluno / area medica) ──────────
@app.route('/api/mensagens-modelo', methods=['GET'])
@login_required
def api_mensagens_modelo():
    db = get_db()
    rows = db.execute('SELECT * FROM mensagens_modelo ORDER BY chave').fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/mensagens-modelo/<chave>', methods=['PUT'])
@login_required
def api_mensagens_modelo_editar(chave):
    if current_user.role != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403
    db = get_db()
    if not db.execute('SELECT chave FROM mensagens_modelo WHERE chave=?', (chave,)).fetchone():
        return jsonify({'erro': 'Nao encontrado'}), 404
    d = request.get_json() or {}
    texto = (d.get('texto') or '').strip()
    if not texto:
        return jsonify({'erro': 'Texto e obrigatorio'}), 400
    db.execute('UPDATE mensagens_modelo SET texto=? WHERE chave=?', (texto, chave))
    db.commit()
    return jsonify({'ok': True})


# ── Residentes & Doutorandos ─────────────────────────────────

@app.route('/residentes')
@login_required
def pagina_residentes():
    db = get_db()
    rows = db.execute('SELECT especialidade FROM limite_especialidade ORDER BY especialidade').fetchall()
    especialidades = [r['especialidade'] for r in rows]
    return render_template('residentes.html', especialidades=especialidades)


@app.route('/api/residentes', methods=['GET'])
@login_required
def api_get_residentes():
    db = get_db()
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', ITENS_POR_PAGINA))
    tipo = request.args.get('tipo', '')
    modalidade = request.args.get('modalidade', '')
    especialidade = request.args.get('especialidade', '')
    mes_ano = request.args.get('mes_ano', '')
    status = request.args.get('status', '')
    status_pagamento = request.args.get('status_pagamento', '')
    busca = request.args.get('busca', '').strip()
    ordenar = request.args.get('ordenar', 'mes_ano')

    conds = []
    params = []

    if tipo:
        conds.append('tipo=?'); params.append(tipo)
    if modalidade:
        conds.append('modalidade=?'); params.append(modalidade)
    if especialidade:
        conds.append('especialidade=?'); params.append(especialidade)
    if mes_ano:
        conds.append('mes_ano=?'); params.append(mes_ano)
    if status:
        conds.append('status=?'); params.append(status)
    if status_pagamento:
        conds.append('status_pagamento=?'); params.append(status_pagamento)
    if busca:
        conds.append('(nome LIKE ? OR email LIKE ? OR cpf LIKE ? OR telefone LIKE ? OR instituicao_origem LIKE ? OR especialidade LIKE ? OR observacao LIKE ?)')
        b = f'%{busca}%'
        params.extend([b, b, b, b, b, b, b])

    ORDENACOES = {
        'recentes': 'created_at DESC, id DESC',
        'nome': 'nome COLLATE NOCASE',
        'mes_ano': 'mes_ano DESC, nome',
    }
    order_sql = ORDENACOES.get(ordenar, ORDENACOES['mes_ano'])

    where = ('WHERE ' + ' AND '.join(conds)) if conds else ''
    total = db.execute(f'SELECT COUNT(*) FROM residentes {where}', params).fetchone()[0]
    offset = (page - 1) * per_page
    rows = db.execute(
        f'''SELECT residentes.*,
                CAST(julianday('now') - julianday(
                    COALESCE(
                        (SELECT MAX(h.ts) FROM historico_residentes h WHERE h.residente_id = residentes.id),
                        residentes.updated_at, residentes.created_at
                    )
                ) AS INTEGER) as dias_no_status
            FROM residentes {where} ORDER BY {order_sql} LIMIT ? OFFSET ?''',
        params + [per_page, offset]
    ).fetchall()

    return jsonify({
        'data': [dict(r) for r in rows],
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': max(1, (total + per_page - 1) // per_page),
    })


@app.route('/api/residentes', methods=['POST'])
@login_required
def api_create_residente():
    db = get_db()
    d = request.get_json()
    if not d.get('nome') or not d.get('especialidade') or not d.get('mes_ano'):
        return jsonify({'erro': 'Nome, especialidade e mes_ano sao obrigatorios'}), 400
    cur = db.execute('''
        INSERT INTO residentes
            (nome, email, telefone, cpf, tipo, modalidade, especialidade, subespecialidade,
             instituicao_origem, programa_ano, mes_ano, inicio, termino, status,
             valor, forma_pagamento, status_pagamento, comprovante_pagamento, observacao,
             data_inscricao, periodo_desejado, mes_desejado)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (
        d.get('nome'), d.get('email'), d.get('telefone'), d.get('cpf'),
        d.get('tipo', 'Residente'), d.get('modalidade', 'Optativo'),
        d.get('especialidade'), d.get('subespecialidade'),
        d.get('instituicao_origem'), d.get('programa_ano'),
        d.get('mes_ano'), d.get('inicio') or None, d.get('termino') or None,
        d.get('status', 'Interessado'),
        d.get('valor') or None, d.get('forma_pagamento'), d.get('status_pagamento', 'Pendente'),
        d.get('comprovante_pagamento'), d.get('observacao'),
        d.get('data_inscricao'), d.get('periodo_desejado'), d.get('mes_desejado'),
    ))
    db.commit()
    return jsonify({'id': cur.lastrowid}), 201


@app.route('/api/residentes/<int:rid>', methods=['PUT'])
@login_required
def api_update_residente(rid):
    db = get_db()
    if not db.execute('SELECT id FROM residentes WHERE id=?', (rid,)).fetchone():
        return jsonify({'erro': 'Nao encontrado'}), 404
    d = request.get_json()
    db.execute('''
        UPDATE residentes SET
            nome=?, email=?, telefone=?, cpf=?, tipo=?, modalidade=?,
            especialidade=?, subespecialidade=?, instituicao_origem=?, programa_ano=?,
            mes_ano=?, inicio=?, termino=?, status=?,
            valor=?, forma_pagamento=?, status_pagamento=?, comprovante_pagamento=?,
            observacao=?, data_inscricao=?, periodo_desejado=?, mes_desejado=?,
            updated_at=CURRENT_TIMESTAMP
        WHERE id=?
    ''', (
        d.get('nome'), d.get('email'), d.get('telefone'), d.get('cpf'),
        d.get('tipo', 'Residente'), d.get('modalidade', 'Optativo'),
        d.get('especialidade'), d.get('subespecialidade'),
        d.get('instituicao_origem'), d.get('programa_ano'),
        d.get('mes_ano'), d.get('inicio') or None, d.get('termino') or None,
        d.get('status', 'Interessado'),
        d.get('valor') or None, d.get('forma_pagamento'), d.get('status_pagamento', 'Pendente'),
        d.get('comprovante_pagamento'), d.get('observacao'),
        d.get('data_inscricao'), d.get('periodo_desejado'), d.get('mes_desejado'), rid,
    ))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/residentes/<int:rid>/historico', methods=['GET'])
@login_required
def api_historico_residente(rid):
    db = get_db()
    rows = db.execute(
        'SELECT * FROM historico_residentes WHERE residente_id=? ORDER BY ts',
        (rid,)
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/residentes/<int:rid>/avancar', methods=['POST'])
@login_required
def api_avancar_residente(rid):
    db = get_db()
    row = db.execute('SELECT * FROM residentes WHERE id=?', (rid,)).fetchone()
    if not row:
        return jsonify({'erro': 'Nao encontrado'}), 404
    d = request.get_json()
    novo_status = d.get('status', row['status'])
    obs = d.get('observacao', '')
    responsavel = current_user.nome if current_user.is_authenticated else 'Sistema'

    db.execute('UPDATE residentes SET status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
               (novo_status, rid))
    db.execute(
        'INSERT INTO historico_residentes (residente_id, status, observacao, responsavel) VALUES (?,?,?,?)',
        (rid, novo_status, obs, responsavel)
    )
    db.commit()
    return jsonify({'status': novo_status})


@app.route('/api/residentes/<int:rid>/pdf')
@login_required
def api_residente_pdf(rid):
    db = get_db()
    row = db.execute('SELECT * FROM residentes WHERE id=?', (rid,)).fetchone()
    if not row:
        return jsonify({'erro': 'Nao encontrado'}), 404

    historico = db.execute(
        'SELECT * FROM historico_residentes WHERE residente_id=? ORDER BY ts', (rid,)
    ).fetchall()

    html = render_template('pdf_ficha_residente.html', r=row, historico=historico,
                            data_geracao=datetime.now().strftime('%d/%m/%Y %H:%M'))
    try:
        pdf_bytes = HTML(string=html).write_pdf()
        nome_arquivo = row['nome'].replace(' ', '_')[:40]
        return Response(pdf_bytes, mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename=ficha_{nome_arquivo}.pdf'})
    except Exception:
        return html


@app.route('/api/residentes/<int:rid>', methods=['DELETE'])
@login_required
def api_delete_residente(rid):
    db = get_db()
    if not db.execute('SELECT id FROM residentes WHERE id=?', (rid,)).fetchone():
        return jsonify({'erro': 'Nao encontrado'}), 404
    db.execute('DELETE FROM residentes WHERE id=?', (rid,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/residentes/exportar-csv', methods=['GET'])
@login_required
def api_exportar_residentes_csv():
    db = get_db()
    sep = request.args.get('separador', ';')
    if sep not in (';', '\t', ','):
        sep = ';'

    tipo = request.args.get('tipo', '')
    modalidade = request.args.get('modalidade', '')
    especialidade = request.args.get('especialidade', '')
    mes_ano = request.args.get('mes_ano', '')
    status = request.args.get('status', '')
    status_pagamento = request.args.get('status_pagamento', '')
    busca = request.args.get('busca', '').strip()

    conds, params = [], []
    if tipo:
        conds.append('tipo=?'); params.append(tipo)
    if modalidade:
        conds.append('modalidade=?'); params.append(modalidade)
    if especialidade:
        conds.append('especialidade=?'); params.append(especialidade)
    if mes_ano:
        conds.append('mes_ano=?'); params.append(mes_ano)
    if status:
        conds.append('status=?'); params.append(status)
    if status_pagamento:
        conds.append('status_pagamento=?'); params.append(status_pagamento)
    if busca:
        conds.append('(nome LIKE ? OR email LIKE ? OR especialidade LIKE ?)')
        b = f'%{busca}%'; params.extend([b, b, b])

    where = ('WHERE ' + ' AND '.join(conds)) if conds else ''
    rows = db.execute(
        f'SELECT * FROM residentes {where} ORDER BY mes_ano DESC, nome', params
    ).fetchall()

    def fmt(v):
        if v is None: return ''
        return str(v).replace(sep, ' ')

    header = ['ID','Nome','Email','Telefone','CPF','Tipo','Modalidade','Especialidade',
              'Subespecialidade','Instituicao','Programa/Ano','Mes/Ano','Inicio','Termino',
              'Status','Valor','Forma Pagamento','Status Pagamento','Comprovante','Observacao',
              'Data Inscricao','Periodo Desejado','Mes Desejado']
    lines = [sep.join(header)]
    for r in rows:
        lines.append(sep.join([
            fmt(r['id']), fmt(r['nome']), fmt(r['email']), fmt(r['telefone']), fmt(r['cpf']),
            fmt(r['tipo']), fmt(r['modalidade']), fmt(r['especialidade']),
            fmt(r['subespecialidade']), fmt(r['instituicao_origem']), fmt(r['programa_ano']),
            fmt(r['mes_ano']), fmt(r['inicio']), fmt(r['termino']),
            fmt(r['status']), fmt(r['valor']), fmt(r['forma_pagamento']),
            fmt(r['status_pagamento']), fmt(r['comprovante_pagamento']), fmt(r['observacao']),
            fmt(r['data_inscricao']), fmt(r['periodo_desejado']), fmt(r['mes_desejado']),
        ]))
    csv_bytes = ('﻿' + '\r\n'.join(lines)).encode('utf-8')
    return Response(csv_bytes, mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=residentes.csv'})


_COL_IMPORT_RESIDENTE = {
    'nome': 'nome', 'nome completo': 'nome',
    'e-mail': 'email', 'email': 'email',
    'whatsapp': 'telefone', 'telefone': 'telefone', 'celular': 'telefone',
    'especialidade': 'especialidade', 'especialidade desejada': 'especialidade',
    'mês desejado': 'mes_texto', 'mes desejado': 'mes_texto',
    'mês(es) desejado(s)': 'mes_texto', 'mes(es) desejado(s)': 'mes_texto',
    'mês desejado 1': 'mes_texto', 'mes desejado 1': 'mes_texto',
    'instituição de origem': 'instituicao_origem', 'instituicao de origem': 'instituicao_origem',
    'instituição': 'instituicao_origem', 'instituicao': 'instituicao_origem',
    'programa e ano (cursando)': 'programa_ano', 'programa/ano': 'programa_ano',
    'programa': 'programa_ano', 'status': 'programa_ano',
    'observação': 'observacao', 'observacao': 'observacao',
    'período desejado (xx/xx/xxxx)': 'periodo_desejado', 'periodo desejado (xx/xx/xxxx)': 'periodo_desejado',
    'mês desejado 2': 'mes_desejado2', 'mes desejado 2': 'mes_desejado2',
    'hora de conclusão': 'data_inscricao', 'hora de conclusao': 'data_inscricao',
    'hora de início': 'data_inscricao', 'hora de inicio': 'data_inscricao',
    'carimbo de data/hora': 'data_inscricao', 'timestamp': 'data_inscricao',
}

_MESES_PT_IMPORT = {
    'janeiro': '01', 'fevereiro': '02', 'março': '03', 'marco': '03',
    'abril': '04', 'maio': '05', 'junho': '06', 'julho': '07',
    'agosto': '08', 'setembro': '09', 'outubro': '10',
    'novembro': '11', 'dezembro': '12',
}


def _parse_mes_texto(texto):
    """Tenta extrair YYYY-MM de um texto livre de mês."""
    if not texto:
        return None
    t = str(texto).lower().strip()
    import re as _re
    # "janeiro 2026", "janeiro/2026", "janeiro de 2026"
    for nome, num in _MESES_PT_IMPORT.items():
        if nome in t:
            m = _re.search(r'(20\d{2})', t)
            ano = m.group(1) if m else '2026'
            return f'{ano}-{num}'
    # "01/2026" ou "01/26"
    m = _re.search(r'(\d{1,2})[/-](\d{2,4})', t)
    if m:
        mes = m.group(1).zfill(2)
        ano = m.group(2)
        if len(ano) == 2:
            ano = '20' + ano
        if 1 <= int(mes) <= 12:
            return f'{ano}-{mes}'
    return None


def _parse_inscricoes_residentes_xlsx(wb, tipo_default='Residente'):
    """Parser para planilhas de inscrição de residentes/doutorandos."""
    import re as _re
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        col_map = {}
        header_found = False
        for row in ws.iter_rows(values_only=True):
            if not any(c for c in row if c is not None):
                continue
            if not header_found:
                candidate = {}
                for j, cell in enumerate(row):
                    if cell is None:
                        continue
                    key = str(cell).strip().lower()
                    if key in _COL_IMPORT_RESIDENTE:
                        candidate[j] = _COL_IMPORT_RESIDENTE[key]
                if 'nome' in candidate.values():
                    col_map = candidate
                    header_found = True
                continue

            rec = {}
            for j, campo in col_map.items():
                val = row[j] if j < len(row) else None
                rec[campo] = _xclean(val)

            nome = rec.get('nome', '')
            if not nome or nome.lower() in ('nome', 'nome completo', ''):
                continue

            especialidade = rec.get('especialidade', '') or ''
            if not especialidade:
                continue

            # mes_ano: tenta campo mes_texto (mês desejado 1), depois periodo_desejado
            mes_texto = rec.get('mes_texto', '') or rec.get('mes_desejado', '') or ''
            mes_ano = _parse_mes_texto(mes_texto)
            if not mes_ano:
                mes_ano = _parse_mes_texto(rec.get('periodo_desejado', '') or '')
            if not mes_ano:
                mes_ano = '2026-01'  # fallback — usuário corrige

            # data_inscricao: guarda como texto para não perder granularidade
            data_inscricao = rec.get('data_inscricao', '') or ''
            if hasattr(data_inscricao, 'strftime'):
                data_inscricao = data_inscricao.strftime('%d/%m/%Y %H:%M')
            else:
                data_inscricao = str(data_inscricao).strip() if data_inscricao else ''

            records.append({
                'nome': nome,
                'email': rec.get('email', ''),
                'telefone': rec.get('telefone', ''),
                'especialidade': especialidade,
                'instituicao_origem': rec.get('instituicao_origem', ''),
                'programa_ano': rec.get('programa_ano', ''),
                'mes_ano': mes_ano,
                'observacao': rec.get('observacao', ''),
                'data_inscricao': data_inscricao,
                'periodo_desejado': rec.get('periodo_desejado', '') or '',
                'mes_desejado': mes_texto,
                'tipo': tipo_default,
                'modalidade': 'Optativo',
                'status': 'Interessado',
                'status_pagamento': 'Pendente',
            })
    return records


@app.route('/api/residentes/importar-excel', methods=['POST'])
@login_required
def api_importar_residentes_excel():
    import openpyxl, io
    confirmar = request.args.get('confirmar', '0') == '1'
    tipo_default = request.args.get('tipo', 'Residente')
    if tipo_default not in TIPOS_RESIDENTE:
        tipo_default = 'Residente'

    f = request.files.get('file')
    if not f:
        return jsonify({'erro': 'Nenhum arquivo enviado'}), 400

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return jsonify({'erro': f'Erro ao ler arquivo: {e}'}), 400

    todos = _parse_inscricoes_residentes_xlsx(wb, tipo_default)

    db = get_db()
    existentes = set()
    for r in db.execute('SELECT nome, especialidade, mes_ano FROM residentes').fetchall():
        existentes.add((r['nome'].lower(), r['especialidade'].lower(), r['mes_ano']))

    novos, duplicados = [], []
    for rec in todos:
        chave = (rec['nome'].lower(), rec['especialidade'].lower(), rec['mes_ano'])
        if chave in existentes:
            duplicados.append(rec)
        else:
            novos.append(rec)
            existentes.add(chave)

    if confirmar:
        for rec in novos:
            db.execute('''
                INSERT INTO residentes
                    (nome, email, telefone, tipo, modalidade, especialidade,
                     instituicao_origem, programa_ano, mes_ano, status,
                     status_pagamento, observacao,
                     data_inscricao, periodo_desejado, mes_desejado)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (rec['nome'], rec.get('email',''), rec.get('telefone',''),
                  rec['tipo'], rec.get('modalidade','Optativo'), rec['especialidade'],
                  rec.get('instituicao_origem',''), rec.get('programa_ano',''), rec['mes_ano'],
                  rec['status'], rec['status_pagamento'], rec.get('observacao',''),
                  rec.get('data_inscricao',''), rec.get('periodo_desejado',''), rec.get('mes_desejado','')))
        db.commit()

    return jsonify({
        'total_planilha': len(todos),
        'novos': len(novos),
        'duplicados': len(duplicados),
        'preview': novos[:30],
    })


# ── Run ───────────────────────────────────────────────────────
if __name__ == '__main__':
    if not os.path.exists(DB_PATH):
        init_db()
    else:
        # Migrate: add new columns if they don't exist
        db = sqlite3.connect(DB_PATH)
        cols = [r[1] for r in db.execute('PRAGMA table_info(estagios)').fetchall()]
        if 'cpf' not in cols:
            db.execute('ALTER TABLE estagios ADD COLUMN cpf TEXT')
        if 'forma_pagamento' not in cols:
            db.execute('ALTER TABLE estagios ADD COLUMN forma_pagamento TEXT')
        if 'status_pagamento' not in cols:
            db.execute("ALTER TABLE estagios ADD COLUMN status_pagamento TEXT DEFAULT 'Pendente'")
        if 'comprovante_pagamento' not in cols:
            db.execute('ALTER TABLE estagios ADD COLUMN comprovante_pagamento TEXT')
        if 'inicio' not in cols:
            db.execute('ALTER TABLE estagios ADD COLUMN inicio DATE')
        if 'comprovante_estagio' not in cols:
            db.execute('ALTER TABLE estagios ADD COLUMN comprovante_estagio INTEGER DEFAULT 0')
        if 'carga_horaria' not in cols:
            db.execute('ALTER TABLE estagios ADD COLUMN carga_horaria INTEGER')
        # Migrar etapa 7 (antigo Concluido) para etapa 8 (novo Concluido)
        # pois etapa 7 agora é Comprovante recebido
        db.execute('UPDATE estagios SET etapa=8 WHERE etapa=7')
        # Make cracha non-unique (allow duplicates for empty/0 values)
        try:
           db.execute('DROP INDEX IF EXISTS idx_estagios_cracha')
        except Exception:
           pass
        # Migrate usuarios table columns
        u_cols = [r[1] for r in db.execute('PRAGMA table_info(usuarios)').fetchall()]
        if u_cols and 'last_login' not in u_cols:
            db.execute('ALTER TABLE usuarios ADD COLUMN last_login DATETIME')
        # Create new tables if not exist
        db.execute('''CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            nome TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            last_login DATETIME
        )''')
        db.execute('''CREATE TABLE IF NOT EXISTS notificacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            estagio_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            mensagem TEXT NOT NULL,
            email_destino TEXT,
            enviado INTEGER DEFAULT 0,
            ts DATETIME DEFAULT CURRENT_TIMESTAMP
        )''')
        db.execute('''CREATE TABLE IF NOT EXISTS limite_especialidade (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            especialidade TEXT UNIQUE NOT NULL COLLATE NOCASE,
            limite_semanal INTEGER NOT NULL DEFAULT 1
        )''')
        # Seed limites if table is empty
        if db.execute('SELECT COUNT(*) FROM limite_especialidade').fetchone()[0] == 0:
            limites_seed = [
                ('Anestesiologia', 4), ('Cardiologia', 4), ('Cirurgia cardiovascular', 2),
                ('Cirurgia de aparelho digestivo', 3), ('Cirurgia de cabeca e pescoco', 3),
                ('Cirurgia geral', 8), ('Cirurgia oncologica', 3), ('Cirurgia pediatrica', 4),
                ('Cirurgia plastica', 3), ('Cirurgia toracica', 2), ('Clinica medica', 8),
                ('Emergencia adulta', 8), ('Emergencia cardiologica', 3), ('Gastroenterologia', 2),
                ('Geriatria', 3), ('Mastologia', 3), ('Neurologia pediatrica', 2),
                ('Neurocirurgia', 4), ('Oncologia clinica', 4), ('Otorrino', 2),
                ('Terapia intensiva', 2),
            ]
            db.executemany('INSERT OR IGNORE INTO limite_especialidade (especialidade, limite_semanal) VALUES (?,?)', limites_seed)
        # Migrar tabela residentes
        db.execute('''CREATE TABLE IF NOT EXISTS residentes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            email TEXT,
            telefone TEXT,
            cpf TEXT,
            tipo TEXT NOT NULL DEFAULT 'Residente',
            modalidade TEXT NOT NULL DEFAULT 'Optativo',
            especialidade TEXT NOT NULL,
            subespecialidade TEXT,
            instituicao_origem TEXT,
            programa_ano TEXT,
            mes_ano TEXT NOT NULL,
            inicio DATE,
            termino DATE,
            status TEXT NOT NULL DEFAULT 'Interessado',
            valor REAL,
            forma_pagamento TEXT,
            status_pagamento TEXT DEFAULT 'Pendente',
            comprovante_pagamento TEXT,
            observacao TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )''')
        res_cols = [r[1] for r in db.execute('PRAGMA table_info(residentes)').fetchall()]
        for col, defn in [('subespecialidade', 'TEXT'), ('programa_ano', 'TEXT'),
                          ('instituicao_origem', 'TEXT'), ('cpf', 'TEXT'),
                          ('valor', 'REAL'), ('forma_pagamento', 'TEXT'),
                          ('status_pagamento', "TEXT DEFAULT 'Pendente'"),
                          ('comprovante_pagamento', 'TEXT')]:
            if res_cols and col not in res_cols:
                db.execute(f'ALTER TABLE residentes ADD COLUMN {col} {defn}')

        # Contatos da area medica + modelos de mensagem (botoes WhatsApp de Residentes)
        db.execute('''CREATE TABLE IF NOT EXISTS area_medica (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            especialidade TEXT NOT NULL,
            nome TEXT,
            celular TEXT,
            email TEXT,
            obs_internato TEXT,
            obs_residencia TEXT
        )''')
        db.execute('''CREATE TABLE IF NOT EXISTS mensagens_modelo (
            chave TEXT PRIMARY KEY,
            titulo TEXT NOT NULL,
            texto TEXT NOT NULL,
            placeholders TEXT
        )''')
        db.executemany(
            'INSERT OR IGNORE INTO mensagens_modelo (chave, titulo, texto, placeholders) VALUES (?,?,?,?)',
            MENSAGENS_MODELO_SEED
        )
        # Seed inicial de area_medica a partir do JSON extraido do documento,
        # apenas se a tabela ainda estiver vazia (edicoes ficam so no banco depois disso)
        if db.execute('SELECT COUNT(*) FROM area_medica').fetchone()[0] == 0:
            am_json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'area_medica.json')
            if os.path.exists(am_json_path):
                with open(am_json_path, encoding='utf-8') as f:
                    contatos = json.load(f)
                db.executemany(
                    '''INSERT INTO area_medica (especialidade, nome, celular, email, obs_internato, obs_residencia)
                       VALUES (?,?,?,?,?,?)''',
                    [(c.get('especialidade'), c.get('nome'), c.get('celular'), c.get('email'),
                      c.get('obs_internato'), c.get('obs_residencia')) for c in contatos]
                )

        # Seed default admin if no users exist
        user_count = db.execute('SELECT COUNT(*) FROM usuarios').fetchone()[0]
        if user_count == 0:
            db.execute("INSERT INTO usuarios (username, password_hash, nome, role) VALUES (?, ?, ?, ?)",
                       ('admin', hash_password('admin'), 'Administrador', 'admin'))
            db.execute("INSERT INTO usuarios (username, password_hash, nome, role) VALUES (?, ?, ?, ?)",
                       ('user', hash_password('user'), 'Usuario', 'user'))
        db.commit()
        db.close()

    app.run(host='0.0.0.0', port=5000, debug=True)
